import os
from openpyxl import Workbook, load_workbook
from tkinter import *
from tkinter import ttk, messagebox, simpledialog

# ---------------- RUTAS ---------------- #
if not os.path.exists("datos"):
    os.makedirs("datos")
archivo_excel = "datos/Ventas.xlsx"

# ---------------- INICIALIZAR ARCHIVO ---------------- #
if not os.path.exists(archivo_excel):
    wb = Workbook()
    wb.create_sheet("Productos", 0)
    wb.create_sheet("Clientes", 1)
    wb.create_sheet("Ventas", 2)
    
    ws = wb["Productos"]
    ws.append(["Código","Nombre","Existencia","Proveedor","Precio"])
    
    ws = wb["Clientes"]
    ws.append(["Código","Nombre","Dirección"])
    
    ws = wb["Ventas"]
    ws.append(["Código Producto","Código Cliente","Cantidad","Total"])
    
    wb.save(archivo_excel)

# ---------------- FUNCIONES EXCEL ---------------- #
def cargar_hoja(nombre_hoja):
    wb = load_workbook(archivo_excel)
    ws = wb[nombre_hoja]
    return [list(row) for row in ws.iter_rows(min_row=2, values_only=True)]

def guardar_hoja(nombre_hoja, datos):
    wb = load_workbook(archivo_excel)
    ws = wb[nombre_hoja]
    ws.delete_rows(2, ws.max_row)
    for fila in datos:
        ws.append(fila)
    wb.save(archivo_excel)

# ---------------- FUNCIONES PRODUCTOS ---------------- #
def listar_productos():
    for i in tabla_productos.get_children():
        tabla_productos.delete(i)
    for p in cargar_hoja("Productos"):
        tabla_productos.insert("", END, values=p)

def crear_producto():
    codigo = ent_codigo_prod.get()
    nombre = ent_nombre_prod.get()
    existencia = ent_existencia.get()
    proveedor = ent_proveedor.get()
    precio = ent_precio.get()
    if not codigo or not nombre or not existencia or not precio:
        messagebox.showerror("Error", "Complete todos los campos")
        return
    productos = cargar_hoja("Productos")
    productos.append([codigo, nombre, int(existencia), proveedor, float(precio)])
    guardar_hoja("Productos", productos)
    listar_productos()

def actualizar_producto():
    productos = cargar_hoja("Productos")
    codigo = ent_codigo_prod.get()
    for p in productos:
        if p[0] == codigo:
            p[1] = ent_nombre_prod.get()
            p[2] = int(ent_existencia.get())
            p[3] = ent_proveedor.get()
            p[4] = float(ent_precio.get())
            guardar_hoja("Productos", productos)
            listar_productos()
            return
    messagebox.showerror("Error", "Producto no encontrado")

def eliminar_producto():
    productos = cargar_hoja("Productos")
    codigo = ent_codigo_prod.get()
    productos = [p for p in productos if p[0] != codigo]
    guardar_hoja("Productos", productos)
    listar_productos()

# ---------------- FUNCIONES CLIENTES ---------------- #
def listar_clientes():
    for i in tabla_clientes.get_children():
        tabla_clientes.delete(i)
    for c in cargar_hoja("Clientes"):
        tabla_clientes.insert("", END, values=c)

def crear_cliente():
    codigo = ent_codigo_cli.get()
    nombre = ent_nombre_cli.get()
    direccion = ent_direccion_cli.get()
    if not codigo or not nombre:
        messagebox.showerror("Error", "Complete todos los campos")
        return
    clientes = cargar_hoja("Clientes")
    clientes.append([codigo, nombre, direccion])
    guardar_hoja("Clientes", clientes)
    listar_clientes()

def actualizar_cliente():
    clientes = cargar_hoja("Clientes")
    codigo = ent_codigo_cli.get()
    for c in clientes:
        if c[0] == codigo:
            c[1] = ent_nombre_cli.get()
            c[2] = ent_direccion_cli.get()
            guardar_hoja("Clientes", clientes)
            listar_clientes()
            return
    messagebox.showerror("Error", "Cliente no encontrado")

def eliminar_cliente():
    clientes = cargar_hoja("Clientes")
    codigo = ent_codigo_cli.get()
    clientes = [c for c in clientes if c[0] != codigo]
    guardar_hoja("Clientes", clientes)
    listar_clientes()

# ---------------- FUNCIONES VENTAS ---------------- #
def listar_ventas():
    for i in tabla_ventas.get_children():
        tabla_ventas.delete(i)
    for v in cargar_hoja("Ventas"):
        tabla_ventas.insert("", END, values=v)

def crear_venta():
    cod_prod = ent_cod_prod_venta.get()
    cod_cli = ent_cod_cli_venta.get()
    cant = ent_cantidad_venta.get()
    if not cod_prod or not cod_cli or not cant:
        messagebox.showerror("Error", "Complete todos los campos")
        return
    cant = int(cant)
    productos = cargar_hoja("Productos")
    ventas = cargar_hoja("Ventas")
    prod = next((p for p in productos if p[0] == cod_prod), None)
    if not prod:
        messagebox.showerror("Error", "Producto no encontrado")
        return
    if prod[2] < cant:
        messagebox.showerror("Error", "No hay suficiente existencia")
        return
    total = cant * prod[4]
    prod[2] -= cant
    ventas.append([cod_prod, cod_cli, cant, total])
    guardar_hoja("Productos", productos)
    guardar_hoja("Ventas", ventas)
    listar_productos()
    listar_ventas()

def anular_venta():
    ventas = cargar_hoja("Ventas")
    cod_prod = simpledialog.askstring("Anular Venta", "Ingrese código de producto de la venta a anular")
    cod_cli = simpledialog.askstring("Anular Venta", "Ingrese código de cliente de la venta a anular")
    nueva_ventas = []
    productos = cargar_hoja("Productos")
    for v in ventas:
        if v[0] == cod_prod and v[1] == cod_cli:
            prod = next((p for p in productos if p[0]==v[0]), None)
            if prod:
                prod[2] += v[2]
        else:
            nueva_ventas.append(v)
    guardar_hoja("Ventas", nueva_ventas)
    guardar_hoja("Productos", productos)
    listar_productos()
    listar_ventas()

# ---------------- FUNCIONES REPORTES ---------------- #
def reporte_ventas_por_cliente():
    ventas = cargar_hoja("Ventas")
    cod_cliente = simpledialog.askstring("Reporte", "Ingrese código del cliente")
    ventas_cli = [v for v in ventas if v[1]==cod_cliente]
    wb = Workbook()
    ws = wb.active
    ws.title = f"Ventas_{cod_cliente}"
    ws.append(["Código Producto","Código Cliente","Cantidad","Total"])
    for v in ventas_cli:
        ws.append(v)
    ruta = f"datos/Reporte_Ventas_Cliente_{cod_cliente}.xlsx"
    wb.save(ruta)
    messagebox.showinfo("Reporte", f"Reporte generado: {ruta}")

def reporte_ventas_por_producto():
    ventas = cargar_hoja("Ventas")
    cod_prod = simpledialog.askstring("Reporte", "Ingrese código del producto")
    ventas_prod = [v for v in ventas if v[0]==cod_prod]
    wb = Workbook()
    ws = wb.active
    ws.title = f"Ventas_{cod_prod}"
    ws.append(["Código Producto","Código Cliente","Cantidad","Total"])
    for v in ventas_prod:
        ws.append(v)
    ruta = f"datos/Reporte_Ventas_Producto_{cod_prod}.xlsx"
    wb.save(ruta)
    messagebox.showinfo("Reporte", f"Reporte generado: {ruta}")

# ---------------- INTERFAZ GRÁFICA ---------------- #
ventana = Tk()
ventana.title("Sistema de Ventas")
ventana.geometry("1000x650")
ventana.config(bg="#e9ecef")

Label(ventana, text="Sistema de Gestión de Ventas", font=("Arial", 18, "bold"), bg="#e9ecef").pack(pady=15)

tabs = ttk.Notebook(ventana)
tabs.pack(fill="both", expand=True, pady=10)

# --- PESTAÑA PRODUCTOS ---
frame_prod = Frame(tabs, bg="#e9ecef")
tabs.add(frame_prod, text="Productos")

Label(frame_prod, text="Código", bg="#e9ecef").grid(row=0,column=0, padx=5, pady=5)
ent_codigo_prod = Entry(frame_prod); ent_codigo_prod.grid(row=0,column=1, padx=5, pady=5)
Label(frame_prod, text="Nombre", bg="#e9ecef").grid(row=1,column=0, padx=5, pady=5)
ent_nombre_prod = Entry(frame_prod); ent_nombre_prod.grid(row=1,column=1, padx=5, pady=5)
Label(frame_prod, text="Existencia", bg="#e9ecef").grid(row=2,column=0, padx=5, pady=5)
ent_existencia = Entry(frame_prod); ent_existencia.grid(row=2,column=1, padx=5, pady=5)
Label(frame_prod, text="Proveedor", bg="#e9ecef").grid(row=3,column=0, padx=5, pady=5)
ent_proveedor = Entry(frame_prod); ent_proveedor.grid(row=3,column=1, padx=5, pady=5)
Label(frame_prod, text="Precio", bg="#e9ecef").grid(row=4,column=0, padx=5, pady=5)
ent_precio = Entry(frame_prod); ent_precio.grid(row=4,column=1, padx=5, pady=5)

Button(frame_prod, text="Crear", command=crear_producto, width=20, bg="#ced4da").grid(row=5,column=0,pady=5)
Button(frame_prod, text="Actualizar", command=actualizar_producto, width=20, bg="#adb5bd").grid(row=5,column=1,pady=5)
Button(frame_prod, text="Eliminar", command=eliminar_producto, width=20, bg="#adb5bd").grid(row=5,column=2,pady=5)

tabla_productos = ttk.Treeview(frame_prod, columns=("Codigo","Nombre","Existencia","Proveedor","Precio"), show="headings", height=8)
for col in ("Codigo","Nombre","Existencia","Proveedor","Precio"):
    tabla_productos.heading(col, text=col)
tabla_productos.grid(row=6,column=0,columnspan=3,pady=10)
listar_productos()

# --- PESTAÑA CLIENTES ---
frame_cli = Frame(tabs, bg="#e9ecef")
tabs.add(frame_cli, text="Clientes")

Label(frame_cli, text="Código", bg="#e9ecef").grid(row=0,column=0, padx=5, pady=5)
ent_codigo_cli = Entry(frame_cli); ent_codigo_cli.grid(row=0,column=1, padx=5, pady=5)
Label(frame_cli, text="Nombre", bg="#e9ecef").grid(row=1,column=0, padx=5, pady=5)
ent_nombre_cli = Entry(frame_cli); ent_nombre_cli.grid(row=1,column=1, padx=5, pady=5)
Label(frame_cli, text="Dirección", bg="#e9ecef").grid(row=2,column=0, padx=5, pady=5)
ent_direccion_cli = Entry(frame_cli); ent_direccion_cli.grid(row=2,column=1, padx=5, pady=5)

Button(frame_cli, text="Crear", command=crear_cliente, width=20, bg="#ced4da").grid(row=3,column=0,pady=5)
Button(frame_cli, text="Actualizar", command=actualizar_cliente, width=20, bg="#adb5bd").grid(row=3,column=1,pady=5)
Button(frame_cli, text="Eliminar", command=eliminar_cliente, width=20, bg="#adb5bd").grid(row=3,column=2,pady=5)

tabla_clientes = ttk.Treeview(frame_cli, columns=("Codigo","Nombre","Direccion"), show="headings", height=8)
for col in ("Codigo","Nombre","Direccion"):
    tabla_clientes.heading(col, text=col)
tabla_clientes.grid(row=4,column=0,columnspan=3,pady=10)
listar_clientes()

# --- PESTAÑA VENTAS ---
frame_ventas = Frame(tabs, bg="#e9ecef")
tabs.add(frame_ventas, text="Ventas")

Label(frame_ventas, text="Código Producto", bg="#e9ecef").grid(row=0,column=0, padx=5, pady=5)
ent_cod_prod_venta = Entry(frame_ventas); ent_cod_prod_venta.grid(row=0,column=1, padx=5, pady=5)
Label(frame_ventas, text="Código Cliente", bg="#e9ecef").grid(row=1,column=0, padx=5, pady=5)
ent_cod_cli_venta = Entry(frame_ventas); ent_cod_cli_venta.grid(row=1,column=1, padx=5, pady=5)
Label(frame_ventas, text="Cantidad", bg="#e9ecef").grid(row=2,column=0, padx=5, pady=5)
ent_cantidad_venta = Entry(frame_ventas); ent_cantidad_venta.grid(row=2,column=1, padx=5, pady=5)

Button(frame_ventas, text="Registrar Venta", command=crear_venta, width=20, bg="#ced4da").grid(row=3,column=0,pady=5)
Button(frame_ventas, text="Anular Venta", command=anular_venta, width=20, bg="#adb5bd").grid(row=3,column=1,pady=5)

tabla_ventas = ttk.Treeview(frame_ventas, columns=("CodigoProducto","CodigoCliente","Cantidad","Total"), show="headings", height=8)
for col in ("CodigoProducto","CodigoCliente","Cantidad","Total"):
    tabla_ventas.heading(col, text=col)
tabla_ventas.grid(row=4,column=0,columnspan=3,pady=10)
listar_ventas()

# --- PESTAÑA REPORTES ---
frame_reportes = Frame(tabs, bg="#e9ecef")
tabs.add(frame_reportes, text="Reportes")

Button(frame_reportes, text="Ventas por Cliente", command=reporte_ventas_por_cliente, width=25, bg="#ced4da").grid(row=0,column=0,padx=5,pady=10)
Button(frame_reportes, text="Ventas por Producto", command=reporte_ventas_por_producto, width=25, bg="#ced4da").grid(row=0,column=1,padx=5,pady=10)

ventana.mainloop()








